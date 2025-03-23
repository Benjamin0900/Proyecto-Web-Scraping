import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sqlalchemy import create_engine
from datetime import datetime

# Configuración de la conexión a PostgreSQL
db_user = "usr_postbi"
db_password = ""
db_host = "10.200.40.10"  # Ejemplo: 'localhost' o '127.0.0.1'
db_port = "5432"  # Puerto por defecto de PostgreSQL
db_name = "DWHP_WEBSCRAP"
db_table = "ft_precios_compe"

# Crear la conexión usando SQLAlchemy
engine = create_engine(f"postgresql://{db_user}:{db_password}@{db_host}:{db_port}/{db_name}")

ruta_plantilla_base = os.path.join(os.path.dirname(__file__), "plantilla_base.csv")

# Verificar si los archivos existen
for ruta in [ruta_plantilla_base]:
    if not os.path.exists(ruta):
        print(f"El archivo {ruta} no existe. Verifica su existencia.")
        exit()

# Obtener la fecha actual en el formato adecuado
fecha_actual = datetime.now().date()

# Cargar los datos en DataFrames
df_plantilla_base = pd.read_csv(ruta_plantilla_base)

# Definir los ruts
rut_blue = '96.938.840-5'
rut_ch_express='96.756.430-3'
rut_correos_chile='60.503.000-9'
rut_starken='96.794.750-4'
rut_ch_express_emprendedores='99.999.999-9'

# Consulta SQL para extraer solo las columnas necesarias con los filtros
query_blue = """
SELECT domicilio AS domicilio_blue_express, 
       sucursal AS sucursal_blue_express
FROM ft_precios_compe 
WHERE DATE(fecha) = %s 
AND rut = %s
ORDER BY orden_original ASC
"""
query_starken = """
SELECT  domicilio AS domicilio_starken, 
        sucursal AS sucursal_starken,
        aereo_a_domicilio AS aereo_domicilio_starken
FROM ft_precios_compe 
WHERE DATE(fecha) = %s 
AND rut = %s
ORDER BY orden_original ASC
"""
query_correos_chile = """
SELECT  domicilio AS domicilio_y_aereo_correos_chile, 
        sucursal AS sucursal_y_aereo_correos_chile,
        terrestre_a_domicilio AS terrestre_domicilio_correos_chile,
        terrestre_a_sucursal AS terrestre_sucursal_correos_chile
FROM ft_precios_compe 
WHERE rut = %s
AND DATE(fecha) = %s
ORDER BY orden_original ASC
"""
query_ch_express = """
SELECT precio AS unico_chile_express
FROM ft_precios_compe 
WHERE DATE(fecha) = %s 
AND rut = %s
ORDER BY orden_original ASC
"""
query_ch_express_emprendedores = """
SELECT  precio AS unico_chile_express_emprendedores,
        fecha
FROM ft_precios_compe 
WHERE DATE(fecha) = %s 
AND rut = %s
ORDER BY orden_original ASC
"""

# Extraer datos y almacenarlos en un DataFrame
df_plantilla_base_seleccionadas = df_plantilla_base[['ORI', 'DEST', 'Tamaño', 'Ruta', '%']]
df_blue_express_seleccionadas = pd.read_sql(query_blue, engine, params=(fecha_actual, rut_blue))
df_starken_seleccionadas = pd.read_sql(query_starken, engine, params=(fecha_actual, rut_starken))
df_chile_express_seleccionadas = pd.read_sql(query_ch_express, engine, params=(fecha_actual, rut_ch_express))
df_correos_chile_seleccionadas = pd.read_sql(query_correos_chile, engine, params=(rut_correos_chile, fecha_actual))
df_chile_express_emprendedores_seleccionadas = pd.read_sql(query_ch_express_emprendedores, engine, params=(fecha_actual, rut_ch_express_emprendedores))

# Concatenar los DataFrames seleccionados
df_entregable = pd.concat([
    df_plantilla_base_seleccionadas,
    df_blue_express_seleccionadas,
    df_starken_seleccionadas,
    df_chile_express_seleccionadas,
    df_correos_chile_seleccionadas,
    df_chile_express_emprendedores_seleccionadas
], axis=1)

# Crear o sobrescribir el archivo entregable.xlsx
output_folder = os.path.join(os.path.dirname(__file__), "Resultados")
os.makedirs(output_folder, exist_ok=True)
excel_file = os.path.join(os.path.dirname(__file__), "Resultados", "entregable.xlsx")
with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    df_entregable.to_excel(writer, index=False, sheet_name="Precios")

# Cargar el archivo Excel para modificar los estilos
wb = load_workbook(excel_file)
ws = wb["Precios"]  # Nombre de la hoja

# Definir los colores de las celdas con tonos más claros
yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Amarillo claro
blue_fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")  # Azul claro
green_fill = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")  # Verde claro
red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")  # Rojo claro

# Identificar la fila de encabezados
header_row = ws[1]  # Primera fila (títulos de las columnas)

# Aplicar colores a las columnas según las reglas definidas
for cell in header_row:
    if cell.value == "unico_chile_express":
        cell.fill = yellow_fill
    elif cell.value in ["domicilio_blue_express", "sucursal_blue_express"]:
        cell.fill = blue_fill
    elif cell.value in ["domicilio_starken", "sucursal_starken", "aereo_domicilio_starken"]:
        cell.fill = green_fill
    elif cell.value in ["domicilio_y_aereo_correos_chile", "terrestre_domicilio_correos_chile",
                        "sucursal_y_aereo_correos_chile", "terrestre_sucursal_correos_chile"]:
        cell.fill = red_fill

# Guardar el archivo con los cambios
wb.save(excel_file)

print(f"Colores aplicados y archivo {excel_file} actualizado exitosamente.")


#-------------------------------------------------------------------------------------------------------------
