from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, numbers
import pandas as pd
import os

# Configuración inicial
file_path = os.path.join(os.path.dirname(__file__), "Resultados", "entregable.xlsx")
wb = load_workbook(file_path)
if "CH express emprendedores" in wb.sheetnames:
    del wb["CH express emprendedores"]
ws = wb.create_sheet("CH express emprendedores")

# Crear estructura base
estructura = [
    ["", "CH express emprendedores", "", "", ""],
    ["Ruta/DIM", "XS", "S", "M", "L"],
    ["Intra"] + [""] * 4,
    ["Cerca"] + [""] * 4,
    ["Lejos Aéreo"] + [""] * 4,
    ["Lejos Terrestre"] + [""] * 4,
]

for row in estructura:
    ws.append(row)

# Mapeo de competidores y columnas
CONFIG = {
    "Intra": {
        "CH express emprendedores": "unico_chile_express_emprendedores"
    },
    "Cerca": {
        "CH express emprendedores": "unico_chile_express_emprendedores"
    },
    "Lejos Aéreo": {
        "CH express emprendedores": "unico_chile_express_emprendedores"
    },
    "Lejos Terrestre": {
        "CH express emprendedores": None
    }
}
COLUMNAS = {"CH express emprendedores": 2}
COLORES = {   
    "CH express emprendedores": "FFEB9C" # Amarillo
}

# Cálculo principal
df = pd.read_excel(file_path, sheet_name="Precios")
for fila, (ruta, config) in enumerate(CONFIG.items(), start=3):
    ruta_filtro = ruta.split()[0].upper()
    df_filtrado = df[df["Ruta"] == ruta_filtro]
    
    for competidor, columna in config.items():
        if columna is None: 
            valores = {t: "Sin Servicio" for t in ["XS", "S", "M", "L"]}
        else:
            try:
                # Filtrar solo los tamaños relevantes
                mask = df_filtrado["Tamaño"].isin(["XS", "S", "M", "L"])
                df_masked = df_filtrado.loc[mask, ["Tamaño", columna, "%"]]
                
                # Calcular ponderaciones simples
                resultados = (
                    df_masked.groupby("Tamaño", group_keys=False)
                    .apply(lambda x: round((x[columna] * x["%"]).sum() / x["%"].sum(), 0))
                    .reindex(["XS", "S", "M", "L"])
                    .fillna(0)
                    .astype(int)
                )
                valores = resultados.to_dict()
            except KeyError:
                valores = {t: "Sin Servicio" for t in ["XS", "S", "M", "L"]}
        
        # Escribir en Excel
        col_start = COLUMNAS[competidor]
        fill = PatternFill(start_color=COLORES[competidor], end_color=COLORES[competidor], fill_type="solid")
        
        for i, tamaño in enumerate(["XS", "S", "M", "L"]):
            celda = ws.cell(row=fila, column=col_start+i)
            valor = valores.get(tamaño, "Sin Servicio")
            celda.value = valor if valor != 0 else "Sin Servicio"
            celda.fill = fill


# Aplicar formato moneda a la primera tabla
for fila in range(3, 7):  # Filas de Intra, Cerca, Lejos Aéreo, Lejos Terrestre
    for col in range(2, 18):  # Columnas donde están los valores
        celda = ws.cell(row=fila, column=col)
        if isinstance(celda.value, (int, float)):  # Verificar si es número
            celda.number_format = celda.number_format = '"$"#,##0' # Formato moneda

wb.save(file_path)