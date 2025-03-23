from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, numbers
import pandas as pd
import os

# Configuración inicial
file_path = os.path.join(os.path.dirname(__file__), "Resultados", "entregable.xlsx")
wb = load_workbook(file_path)
if "Ponderaciones domicilio" in wb.sheetnames:
    del wb["Ponderaciones domicilio"]
ws = wb.create_sheet("Ponderaciones domicilio")

# Crear estructura base
estructura = [
    ["", "CCH", "", "", "", "Starken", "", "", "", "Blue", "", "", "", "CH express", "", "", ""],
    ["Ruta/DIM", "XS", "S", "M", "L", "XS", "S", "M", "L", "XS", "S", "M", "L", "XS", "S", "M", "L"],
    ["Intra"] + [""] * 16,
    ["Cerca"] + [""] * 16,
    ["Lejos Aéreo"] + [""] * 16,
    ["Lejos Terrestre"] + [""] * 16
]

for row in estructura:
    ws.append(row)

# Mapeo de competidores y columnas
CONFIG = {
    "Intra": {
        "CCH": "domicilio_y_aereo_correos_chile",
        "Starken": "domicilio_starken",
        "Blue": "domicilio_blue_express",
        "CH express": "unico_chile_express"
    },
    "Cerca": {
        "CCH": "domicilio_y_aereo_correos_chile",
        "Starken": "domicilio_starken",
        "Blue": "domicilio_blue_express",
        "CH express": "unico_chile_express"
    },
    "Lejos Aéreo": {
        "CCH": "domicilio_y_aereo_correos_chile",
        "Starken": "aereo_domicilio_starken",
        "Blue": None,
        "CH express": "unico_chile_express"
    },
    "Lejos Terrestre": {
        "CCH": "terrestre_domicilio_correos_chile",
        "Starken": "domicilio_starken",
        "Blue": "domicilio_blue_express",
        "CH express": None
    }
}

COLUMNAS = {"CCH": 2, "Starken": 6, "Blue": 10, "CH express": 14}
COLORES = {
    "CCH": "FFC7CE",    # Rojo claro
    "Starken": "C6EFCE", # Verde claro
    "Blue": "BDD7EE",   # Azul claro
    "CH express": "FFEB9C" # Amarillo
}

# Cálculo principal
df = pd.read_excel(file_path, sheet_name="Precios")
for fila, (ruta, config) in enumerate(CONFIG.items(), start=3):
    ruta_filtro = ruta.split()[0].upper()
    df_filtrado = df[df["Ruta"] == ruta_filtro]
    
    for competidor, columna in config.items():
        if columna is None: 
            valores = {t: "Sin Servicio" for t in ["XS", "S", "M", "L"]}
        else:
            try:
                # Filtrar solo los tamaños relevantes
                mask = df_filtrado["Tamaño"].isin(["XS", "S", "M", "L"])
                df_masked = df_filtrado.loc[mask, ["Tamaño", columna, "%"]]
                
                # Calcular ponderaciones simples
                resultados = (
                    df_masked.groupby("Tamaño", group_keys=False)
                    .apply(lambda x: round((x[columna] * x["%"]).sum() / x["%"].sum(), 0))
                    .reindex(["XS", "S", "M", "L"])
                    .fillna(0)
                    .astype(int)
                )
                valores = resultados.to_dict()
            except KeyError:
                valores = {t: "Sin Servicio" for t in ["XS", "S", "M", "L"]}
        
        # Escribir en Excel
        col_start = COLUMNAS[competidor]
        fill = PatternFill(start_color=COLORES[competidor], end_color=COLORES[competidor], fill_type="solid")
        
        for i, tamaño in enumerate(["XS", "S", "M", "L"]):
            celda = ws.cell(row=fila, column=col_start+i)
            valor = valores.get(tamaño, "Sin Servicio")
            celda.value = valor if valor != 0 else "Sin Servicio"
            celda.fill = fill


# Aplicar formato moneda a la primera tabla
for fila in range(3, 7):  # Filas de Intra, Cerca, Lejos Aéreo, Lejos Terrestre
    for col in range(2, 18):  # Columnas donde están los valores
        celda = ws.cell(row=fila, column=col)
        if isinstance(celda.value, (int, float)):  # Verificar si es número
            celda.number_format = celda.number_format = '"$"#,##0' # Formato moneda


COLORES_c = {
    "CCH": "f14949",    # Rojo claro
    "Starken": "66fe54", # Verde claro
    "Blue": "4d8bff",   # Azul claro
    "CH express": "fef954" # Amarillo
}

# Colorear encabezados
header_fill_cch = PatternFill(start_color=COLORES_c["CCH"], end_color=COLORES_c["CCH"], fill_type="solid")
header_fill_starken = PatternFill(start_color=COLORES_c["Starken"], end_color=COLORES_c["Starken"], fill_type="solid")
header_fill_blue = PatternFill(start_color=COLORES_c["Blue"], end_color=COLORES_c["Blue"], fill_type="solid")
header_fill_ch_express = PatternFill(start_color=COLORES_c["CH express"], end_color=COLORES_c["CH express"], fill_type="solid")

for col in range(2, 6):  # Columnas CCH
    ws.cell(row=1, column=col).fill = header_fill_cch
    
for col in range(6, 10):  # Columnas Starken
    ws.cell(row=1, column=col).fill = header_fill_starken
    
for col in range(10, 14):  # Columnas Blue
    ws.cell(row=1, column=col).fill = header_fill_blue
    
for col in range(14, 18):  # Columnas CH express
    ws.cell(row=1, column=col).fill = header_fill_ch_express

# Ajustes finales
for col in 'BCDEFGHIJKLMNOPQ':
    ws.column_dimensions[col].width = 9


# Definir colores para formato condicional
verde_claro = PatternFill(start_color="cdffc2", end_color="cdffc2", fill_type="solid")
rojo_claro = PatternFill(start_color="ffc2c2", end_color="ffc2c2", fill_type="solid")
gris_claro = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type= "solid")
kk_claro = PatternFill(start_color="f3f0db", end_color="f3f0db", fill_type= "solid")

# Dejar espacio entre las tablas
for _ in range(2):  # Dos filas vacías para separar
    ws.append([""])

# Crear encabezado para la nueva tabla
ws.append(["", "CCH vs Starken","","","","CCH vs Blue","","","","CCH vs CH express"])
ws.append(["Ruta/DIM", "XS", "S", "M", "L",
           "XS", "S", "M", "L",
           "XS", "S", "M", "L"])

# Calcular diferencias y llenar la tabla de comparaciones
start_row = ws.max_row + 1  # Primera fila de la nueva tabla

for fila in range(3, 7):  # Filas de Intra, Cerca, Lejos Aéreo, Lejos Terrestre
    comparacion = [ws.cell(row=fila, column=1).value]  # Ruta/DIM
    
    for col in range(6, 18, 4):  # Columnas de Starken, Blue, CH express
        for i in range(4):  # Para XS, S, M, L
            celda_cch = ws.cell(row=fila, column=2+i)
            celda_otro = ws.cell(row=fila, column=col+i)
            
            if celda_cch.value == "Sin Servicio" or celda_otro.value == "Sin Servicio":
                diferencia = "Sin Servicio"
            elif isinstance(celda_cch.value, int) and isinstance(celda_otro.value, int):
                diferencia = ((celda_otro.value - celda_cch.value) / celda_cch.value) * 100
                diferencia = round(diferencia, 2)
            else:
                diferencia = "Sin Servicio"
            
            comparacion.append(diferencia)
    
    ws.append(comparacion)
contador=0
# Aplicar colores a las diferencias en la tabla inferior
for fila in range(start_row -2, ws.max_row + 1):
    for col in range(2, 14):  # Solo las columnas de diferencias
        celda = ws.cell(row=fila, column=col)

        if contador == 1 or contador == 2 or contador == 3:
            celda.fill= header_fill_starken
            contador+=1
        elif contador == 5 or contador == 6 or contador == 7:
            celda.fill= header_fill_blue
            contador+=1
        elif contador == 9 or contador == 10 or contador == 11:
            contador+=1
            celda.fill=header_fill_ch_express

        if celda.value == "Sin Servicio":
            celda.fill = kk_claro
        elif celda.value == "CCH vs Starken":
            celda.fill = header_fill_starken
            contador+=1
        elif celda.value == "CCH vs Blue":
            celda.fill = header_fill_blue
            contador+=1
        elif celda.value == "CCH vs CH express":
            celda.fill = header_fill_ch_express
            contador+=1

        if isinstance(celda.value, (int, float)):  # Asegurar que es un número
            if celda.value > 0:
                celda.fill = verde_claro
            elif celda.value < 0:
                celda.fill = rojo_claro
            elif celda.value == 0:
                celda.fill= gris_claro

# Aplicar bordes a todas las celdas
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

# Aplicar bordes solo a la primera tabla (filas 1 a 6, columnas 1 a 17)
for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=17):
    for cell in row:
        cell.border = thin_border

# Aplicar bordes solo a la segunda tabla (filas 9 a ws.max_row, columnas 1 a 13)
for row in ws.iter_rows(min_row=9, max_row=ws.max_row, min_col=1, max_col=13):
    for cell in row:
        cell.border = thin_border


wb.save(file_path)